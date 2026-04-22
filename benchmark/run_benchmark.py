"""Benchmark harness: compare labeled ground truth against PhotoSift's
predictions stored in ~/.photosift/photosift.db.

Reads the workbook (Faces / Groups / Sharpness sheets), joins each labeled
row to the PhotoSift `photos` row via raw_path, then computes:

FACES
  - recall, precision, F1 on face-present detection
  - MAE on face count
  - per-slice recall on (small face / profile / occlusion / sunglasses / low-light)
  - per-row error_type (missed / extra / correct) written back to Faces sheet

GROUPS
  - pair-level precision, recall, F1 on "same cluster"
  - Adjusted Rand Index
  - min hamming distance intra- vs inter-cluster per labeled image (written back)
  - near-dup / related tier confusion matrix

SHARPNESS
  - Spearman rank correlation between human 1–10 and PhotoSift sharpness
  - accuracy of a bucketed prediction against human category
  - per-row abs_score_delta written back (after harness-normalized mapping)

Results land on the `Results` sheet and a standalone `benchmark_report.md`
alongside the workbook.

Usage:
    python run_benchmark.py --db ~/.photosift/photosift.db

    # to target a specific shoot set:
    python run_benchmark.py --db ~/.photosift/photosift.db --shoots my_shoot_a my_shoot_b
"""

from __future__ import annotations

import argparse
import itertools
import json
import math
import sqlite3
import statistics
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


HEADER_FILL = PatternFill("solid", start_color="1F2937")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)


# ----- DB access -----


def load_db_rows(db_path: Path, raw_paths: set[str]) -> dict[str, dict]:
    """Pull one row per labeled raw_path. Left-joins faces and groups so
    missing AI rows don't crash the join — we just report None."""
    con = sqlite3.connect(str(db_path))
    con.row_factory = sqlite3.Row
    cur = con.cursor()

    # Batched WHERE IN. SQLite has a param limit, so we chunk.
    out: dict[str, dict] = {}
    paths = list(raw_paths)
    chunk = 400
    for i in range(0, len(paths), chunk):
        batch = paths[i : i + chunk]
        placeholders = ",".join("?" * len(batch))
        cur.execute(
            f"""
            SELECT
                p.id                AS photo_id,
                p.raw_path          AS raw_path,
                p.shoot_id          AS shoot_id,
                p.filename          AS filename,
                p.phash             AS phash,
                p.face_count        AS face_count,
                p.eyes_open_count   AS eyes_open_count,
                p.sharpness_score   AS sharpness_score,
                p.quality_score     AS quality_score,
                p.ai_analyzed_at    AS ai_analyzed_at
            FROM photos p
            WHERE p.raw_path IN ({placeholders})
            """,
            batch,
        )
        for r in cur.fetchall():
            out[r["raw_path"]] = dict(r)

    # Pull faces per matched photo_id.
    photo_ids = [r["photo_id"] for r in out.values()]
    faces_by_photo: dict[int, list[dict]] = defaultdict(list)
    for i in range(0, len(photo_ids), chunk):
        batch = photo_ids[i : i + chunk]
        placeholders = ",".join("?" * len(batch))
        cur.execute(
            f"""
            SELECT photo_id, bbox_x, bbox_y, bbox_w, bbox_h,
                   left_eye_open, right_eye_open,
                   left_eye_sharpness, right_eye_sharpness,
                   detection_confidence
            FROM faces
            WHERE photo_id IN ({placeholders})
            """,
            batch,
        )
        for r in cur.fetchall():
            faces_by_photo[r["photo_id"]].append(dict(r))

    # Pull group memberships per matched photo_id.
    groups_by_photo: dict[int, tuple[int, str]] = {}
    for i in range(0, len(photo_ids), chunk):
        batch = photo_ids[i : i + chunk]
        placeholders = ",".join("?" * len(batch))
        cur.execute(
            f"""
            SELECT gm.photo_id, gm.group_id, g.group_type
            FROM group_members gm
            JOIN groups g ON g.id = gm.group_id
            WHERE gm.photo_id IN ({placeholders})
            """,
            batch,
        )
        for r in cur.fetchall():
            groups_by_photo[r["photo_id"]] = (r["group_id"], r["group_type"])

    for raw, row in out.items():
        row["faces"] = faces_by_photo.get(row["photo_id"], [])
        gm = groups_by_photo.get(row["photo_id"])
        row["group_id_pred"] = gm[0] if gm else None
        row["group_type_pred"] = gm[1] if gm else None

    # Pull thresholds + AI provider status if present.
    cur.execute("SELECT near_dup_threshold, related_threshold FROM settings WHERE id = 1")
    settings = dict(cur.fetchone())

    con.close()
    return {"rows": out, "settings": settings}


# ----- Faces -----


def eval_faces(faces_sheet, db_rows: dict[str, dict]) -> dict:
    header = [c.value for c in faces_sheet[1]]
    idx = {h: i for i, h in enumerate(header)}

    tp = fp = fn = tn = 0
    abs_err_sum = 0
    n = 0
    slice_stats: dict[str, dict[str, int]] = defaultdict(lambda: {"hits": 0, "misses": 0})

    for row in faces_sheet.iter_rows(min_row=2, values_only=False):
        if row[idx["raw_path"]].value is None:
            continue
        unsure = (row[idx["unsure"]].value or "").strip().upper() == "Y"
        if unsure:
            continue
        true_val = row[idx["face_count_true"]].value
        if true_val is None or true_val == "":
            continue  # not yet labeled
        try:
            true_count = int(true_val)
        except (TypeError, ValueError):
            continue

        raw = row[idx["raw_path"]].value
        db = db_rows.get(raw)
        pred_count = db["face_count"] if db and db["face_count"] is not None else None
        confidences = [f["detection_confidence"] for f in (db["faces"] if db else [])]

        # Populate predicted cells.
        row[idx["face_count_pred"]].value = pred_count
        row[idx["max_confidence"]].value = max(confidences) if confidences else None
        row[idx["min_confidence"]].value = min(confidences) if confidences else None

        if pred_count is None:
            # Not analyzed — skip from metrics but mark as unsure so it's visible.
            row[idx["error_type"]].value = "unsure"
            continue

        n += 1
        abs_err = abs(pred_count - true_count)
        abs_err_sum += abs_err
        row[idx["abs_error"]].value = abs_err

        # Binary presence/absence.
        true_has = true_count > 0
        pred_has = pred_count > 0
        if true_has and pred_has:
            tp += 1
        elif true_has and not pred_has:
            fn += 1
        elif pred_has and not true_has:
            fp += 1
        else:
            tn += 1

        if pred_count < true_count:
            row[idx["error_type"]].value = "missed"
        elif pred_count > true_count:
            row[idx["error_type"]].value = "extra"
        else:
            row[idx["error_type"]].value = "correct"

        # Per-slice recall on missed faces.
        for cond in ["has_profile", "has_small_face", "has_occlusion",
                     "has_sunglasses", "has_low_light", "has_motion", "has_backlit"]:
            v = (row[idx[cond]].value or "").strip().upper()
            if v == "Y" and true_has:
                if pred_has and pred_count >= true_count:
                    slice_stats[cond]["hits"] += 1
                else:
                    slice_stats[cond]["misses"] += 1

    recall = tp / (tp + fn) if (tp + fn) else float("nan")
    precision = tp / (tp + fp) if (tp + fp) else float("nan")
    f1 = (
        2 * precision * recall / (precision + recall)
        if (precision + recall and not math.isnan(precision + recall))
        else float("nan")
    )
    mae = abs_err_sum / n if n else float("nan")

    per_slice = {}
    for cond, s in slice_stats.items():
        denom = s["hits"] + s["misses"]
        per_slice[cond] = {
            "recall": s["hits"] / denom if denom else float("nan"),
            "n": denom,
        }

    return {
        "n": n,
        "tp": tp, "fp": fp, "fn": fn, "tn": tn,
        "recall_presence": recall,
        "precision_presence": precision,
        "f1_presence": f1,
        "mae_count": mae,
        "per_slice_recall": per_slice,
    }


# ----- Groups -----


def pair_metrics(true_labels: list[int], pred_labels: list[int | None]) -> dict:
    """Pair-based P/R/F1 where a 'positive' pair is same-cluster in truth.
    Undefined pred (not clustered) is treated as its own singleton so it counts
    toward precision/recall computation."""
    # Replace None with unique sentinels so they can't match anything.
    next_sentinel = -1_000_000
    pred_clean: list[int] = []
    for p in pred_labels:
        if p is None:
            pred_clean.append(next_sentinel)
            next_sentinel -= 1
        else:
            pred_clean.append(p)

    tp = fp = fn = tn = 0
    for i, j in itertools.combinations(range(len(true_labels)), 2):
        same_true = true_labels[i] == true_labels[j]
        same_pred = pred_clean[i] == pred_clean[j]
        if same_true and same_pred:
            tp += 1
        elif same_true and not same_pred:
            fn += 1
        elif same_pred and not same_true:
            fp += 1
        else:
            tn += 1
    prec = tp / (tp + fp) if (tp + fp) else float("nan")
    rec = tp / (tp + fn) if (tp + fn) else float("nan")
    f1 = (
        2 * prec * rec / (prec + rec)
        if (prec + rec and not math.isnan(prec + rec))
        else float("nan")
    )
    return {"pair_precision": prec, "pair_recall": rec, "pair_f1": f1, "pairs_tp": tp, "pairs_fp": fp, "pairs_fn": fn}


def adjusted_rand_index(true_labels: list[int], pred_labels: list[int]) -> float:
    """Standard ARI. Returns nan if we don't have enough samples."""
    n = len(true_labels)
    if n < 2:
        return float("nan")

    def contingency(a, b):
        tbl: dict[tuple[int, int], int] = defaultdict(int)
        for x, y in zip(a, b):
            tbl[(x, y)] += 1
        return tbl

    def comb2(x):
        return x * (x - 1) // 2

    tbl = contingency(true_labels, pred_labels)
    row_sums: dict[int, int] = defaultdict(int)
    col_sums: dict[int, int] = defaultdict(int)
    for (r, c), v in tbl.items():
        row_sums[r] += v
        col_sums[c] += v

    sum_comb_tbl = sum(comb2(v) for v in tbl.values())
    sum_comb_rows = sum(comb2(v) for v in row_sums.values())
    sum_comb_cols = sum(comb2(v) for v in col_sums.values())
    total_comb = comb2(n)

    expected = sum_comb_rows * sum_comb_cols / total_comb if total_comb else 0.0
    max_index = 0.5 * (sum_comb_rows + sum_comb_cols)
    if math.isclose(max_index, expected):
        return 1.0
    return (sum_comb_tbl - expected) / (max_index - expected)


def hamming_distance_hex(a: bytes | None, b: bytes | None) -> int | None:
    if a is None or b is None or len(a) != len(b):
        return None
    d = 0
    for x, y in zip(a, b):
        d += bin(x ^ y).count("1")
    return d


def eval_groups(groups_sheet, db_rows: dict[str, dict]) -> dict:
    header = [c.value for c in groups_sheet[1]]
    idx = {h: i for i, h in enumerate(header)}

    labeled_rows = []
    for row in groups_sheet.iter_rows(min_row=2, values_only=False):
        if row[idx["raw_path"]].value is None:
            continue
        unsure = (row[idx["unsure"]].value or "").strip().upper() == "Y"
        if unsure:
            continue
        cid = row[idx["cluster_id_true"]].value
        if cid in (None, ""):
            continue
        try:
            cid_int = int(cid)
        except (TypeError, ValueError):
            continue
        raw = row[idx["raw_path"]].value
        db = db_rows.get(raw)
        pred = db["group_id_pred"] if db else None
        pred_type = db["group_type_pred"] if db else None

        # Write prediction back.
        row[idx["cluster_id_pred"]].value = pred
        row[idx["group_type_pred"]].value = pred_type

        labeled_rows.append({
            "raw_path": raw,
            "cid_true": cid_int,
            "expected_link": row[idx["expected_link"]].value,
            "burst_type": row[idx["burst_type"]].value,
            "cid_pred": pred,
            "type_pred": pred_type,
            "phash": db["phash"] if db else None,
            "sheet_row": row,
            "idx": idx,
        })

    if len(labeled_rows) < 2:
        return {"n": len(labeled_rows), "note": "need ≥2 labeled rows"}

    # Per-row min hamming intra / inter — we compute this against all OTHER
    # labeled rows, not against the whole library. That's correct: the harness
    # is evaluating the labeled set.
    for r in labeled_rows:
        if r["phash"] is None:
            continue
        intra = []
        inter = []
        for s in labeled_rows:
            if s is r or s["phash"] is None:
                continue
            d = hamming_distance_hex(r["phash"], s["phash"])
            if d is None:
                continue
            if s["cid_true"] == r["cid_true"]:
                intra.append(d)
            else:
                inter.append(d)
        if intra:
            r["sheet_row"][r["idx"]["min_hamming_intra"]].value = min(intra)
        if inter:
            r["sheet_row"][r["idx"]["min_hamming_inter"]].value = min(inter)

    true_labels = [r["cid_true"] for r in labeled_rows]
    pred_labels = [r["cid_pred"] for r in labeled_rows]
    pair = pair_metrics(true_labels, pred_labels)

    # For ARI we need deterministic int labels even for None predicted.
    sentinel = -1_000_000
    pred_ari = []
    for p in pred_labels:
        if p is None:
            pred_ari.append(sentinel)
            sentinel -= 1
        else:
            pred_ari.append(p)
    ari = adjusted_rand_index(true_labels, pred_ari)

    # Tier confusion: expected_link vs group_type_pred.
    expected_counts: dict[tuple[str, str], int] = defaultdict(int)
    for r in labeled_rows:
        exp = r["expected_link"] or "?"
        pred_type = r["type_pred"] or "unclustered"
        expected_counts[(exp, pred_type)] += 1

    return {
        "n": len(labeled_rows),
        "ari": ari,
        **pair,
        "tier_confusion": dict(expected_counts),
    }


# ----- Sharpness -----


def eval_sharpness(sharp_sheet, db_rows: dict[str, dict]) -> dict:
    header = [c.value for c in sharp_sheet[1]]
    idx = {h: i for i, h in enumerate(header)}

    # Collect raw sharpness for percentile mapping *within this benchmark run*.
    vals = []
    rows_with_data = []
    for row in sharp_sheet.iter_rows(min_row=2, values_only=False):
        if row[idx["raw_path"]].value is None:
            continue
        db = db_rows.get(row[idx["raw_path"]].value)
        if not db:
            continue
        sharp = db.get("sharpness_score")
        row[idx["sharpness_raw"]].value = sharp
        row[idx["face_mean_eye_sharp"]].value = (
            statistics.fmean(
                list(
                    itertools.chain.from_iterable(
                        (f["left_eye_sharpness"], f["right_eye_sharpness"]) for f in db["faces"]
                    )
                )
            ) if db["faces"] else None
        )
        row[idx["eyes_open_count"]].value = db.get("eyes_open_count")
        row[idx["quality_score_db"]].value = db.get("quality_score")
        if sharp is not None:
            vals.append(sharp)
            rows_with_data.append((row, db))

    # Percentile rank per row.
    if vals:
        sorted_vals = sorted(vals)
        for row, db in rows_with_data:
            s = db["sharpness_score"]
            # Fraction of sampled values <= s.
            lo, hi = 0, len(sorted_vals)
            while lo < hi:
                mid = (lo + hi) // 2
                if sorted_vals[mid] <= s:
                    lo = mid + 1
                else:
                    hi = mid
            pct = lo / len(sorted_vals) * 100
            row[idx["sharpness_pct"]].value = round(pct, 1)

    # Now compute metrics on unsure=N, labeled rows.
    pairs = []  # (human_1_10, raw_sharpness)
    cat_pairs = []  # (human_category, bucketed_pred)
    for row in sharp_sheet.iter_rows(min_row=2, values_only=False):
        if row[idx["raw_path"]].value is None:
            continue
        unsure = (row[idx["unsure"]].value or "").strip().upper() == "Y"
        if unsure:
            continue
        true_score = row[idx["focus_score_true"]].value
        if true_score in (None, ""):
            continue
        try:
            true_score_f = float(true_score)
        except (TypeError, ValueError):
            continue
        db = db_rows.get(row[idx["raw_path"]].value)
        if not db:
            continue
        pct = row[idx["sharpness_pct"]].value
        if pct is None:
            continue

        # Bucketed prediction: map percentile to one of the human categories
        # using heuristic cutoffs. These are explicit so the user can tune.
        if pct >= 80:
            bucket = "tack"
        elif pct >= 55:
            bucket = "good"
        elif pct >= 30:
            bucket = "acceptable_soft"
        elif pct >= 10:
            bucket = "motion_blur"   # ambiguous — could be OOF or motion; harness marks as motion_blur by convention
        else:
            bucket = "oof"
        row[idx["bucketed_pred"]].value = bucket

        # abs_score_delta: map percentile (0-100) to 1-10 linearly and diff.
        predicted_1_10 = 1 + pct / 100 * 9
        row[idx["abs_score_delta"]].value = round(abs(predicted_1_10 - true_score_f), 2)

        pairs.append((true_score_f, float(db["sharpness_score"])))
        cat_true = row[idx["category_true"]].value
        if cat_true:
            cat_pairs.append((cat_true, bucket))

    if not pairs:
        return {"n": 0, "note": "no labeled rows with PhotoSift sharpness"}

    rho = spearman(pairs)
    mae = sum(abs(a - (1 + percentile(b, [p[1] for p in pairs]) / 100 * 9)) for a, b in pairs) / len(pairs)

    # Category accuracy.
    cat_acc = 0
    if cat_pairs:
        cat_acc = sum(1 for t, p in cat_pairs if t == p) / len(cat_pairs)

    return {
        "n": len(pairs),
        "spearman_rho": rho,
        "mae_1_10": mae,
        "category_accuracy": cat_acc,
        "category_n": len(cat_pairs),
    }


def percentile(x: float, all_vals: list[float]) -> float:
    s = sorted(all_vals)
    lo, hi = 0, len(s)
    while lo < hi:
        mid = (lo + hi) // 2
        if s[mid] <= x:
            lo = mid + 1
        else:
            hi = mid
    return lo / len(s) * 100


def spearman(pairs: list[tuple[float, float]]) -> float:
    """Spearman rank correlation. Handles ties via average ranks."""
    if len(pairs) < 2:
        return float("nan")
    xs = [p[0] for p in pairs]
    ys = [p[1] for p in pairs]
    rx = _avg_ranks(xs)
    ry = _avg_ranks(ys)
    mx = sum(rx) / len(rx)
    my = sum(ry) / len(ry)
    num = sum((a - mx) * (b - my) for a, b in zip(rx, ry))
    den = math.sqrt(sum((a - mx) ** 2 for a in rx) * sum((b - my) ** 2 for b in ry))
    return num / den if den else float("nan")


def _avg_ranks(xs: list[float]) -> list[float]:
    sorted_idx = sorted(range(len(xs)), key=lambda i: xs[i])
    ranks = [0.0] * len(xs)
    i = 0
    while i < len(xs):
        j = i
        while j + 1 < len(xs) and xs[sorted_idx[j + 1]] == xs[sorted_idx[i]]:
            j += 1
        avg = (i + j) / 2 + 1  # ranks start at 1
        for k in range(i, j + 1):
            ranks[sorted_idx[k]] = avg
        i = j + 1
    return ranks


# ----- Output -----


def write_results(wb, results: dict, meta: dict):
    ws = wb["Results"]
    # Clear data rows below header (header at row 4 per build_workbook.py).
    if ws.max_row > 4:
        ws.delete_rows(5, ws.max_row - 4)

    def add(feature: str, metric: str, value, n="", notes=""):
        ws.append([metric, feature, value, n, "", notes])

    # Faces
    f = results["faces"]
    add("faces", "recall_presence", f.get("recall_presence"), f.get("n"))
    add("faces", "precision_presence", f.get("precision_presence"), f.get("n"))
    add("faces", "f1_presence", f.get("f1_presence"), f.get("n"))
    add("faces", "mae_count", f.get("mae_count"), f.get("n"))
    for cond, s in f.get("per_slice_recall", {}).items():
        add("faces", f"recall_slice:{cond}", s["recall"], s["n"])

    # Groups
    g = results["groups"]
    if g.get("n", 0) >= 2:
        add("groups", "adjusted_rand_index", g.get("ari"), g.get("n"))
        add("groups", "pair_precision", g.get("pair_precision"), g.get("n"))
        add("groups", "pair_recall", g.get("pair_recall"), g.get("n"))
        add("groups", "pair_f1", g.get("pair_f1"), g.get("n"))
        for (exp, pred_type), c in g.get("tier_confusion", {}).items():
            add("groups", f"tier_confusion:{exp}->{pred_type}", c)
    else:
        add("groups", "note", g.get("note", "not enough data"), g.get("n"))

    # Sharpness
    s = results["sharpness"]
    if s.get("n"):
        add("sharpness", "spearman_rho", s.get("spearman_rho"), s.get("n"))
        add("sharpness", "mae_1_10", s.get("mae_1_10"), s.get("n"))
        add("sharpness", "category_accuracy", s.get("category_accuracy"), s.get("category_n"))
    else:
        add("sharpness", "note", s.get("note", "not enough data"))

    # Format numeric value column.
    for row in ws.iter_rows(min_row=5, min_col=3, max_col=3):
        for cell in row:
            if isinstance(cell.value, float):
                cell.number_format = "0.000"

    # Write _RunMeta.
    rm = wb["_RunMeta"]
    lookup = {rm.cell(row=r, column=1).value: r for r in range(4, rm.max_row + 1)}
    for k, v in meta.items():
        if k in lookup:
            rm.cell(row=lookup[k], column=2, value=v)


def write_report(results: dict, meta: dict, out_path: Path):
    lines = []
    lines.append("# PhotoSift AI Benchmark — Report")
    lines.append("")
    lines.append(f"- Run at: {meta.get('run_at')}")
    lines.append(f"- DB: `{meta.get('db_path')}`")
    lines.append(f"- Provider: {meta.get('providerStatus', 'unknown')} "
                 f"(eye={meta.get('eyeProviderKind', 'unknown')}, mouth={meta.get('mouthProviderKind', 'unknown')})")
    lines.append("")

    f = results["faces"]
    lines.append("## Face detection")
    lines.append(f"- N labeled: **{f.get('n', 0)}**")
    lines.append(f"- Presence recall: **{fmt(f.get('recall_presence'))}**")
    lines.append(f"- Presence precision: **{fmt(f.get('precision_presence'))}**")
    lines.append(f"- F1: **{fmt(f.get('f1_presence'))}**")
    lines.append(f"- MAE on face count: **{fmt(f.get('mae_count'))}**")
    slices = f.get("per_slice_recall", {})
    if slices:
        lines.append("")
        lines.append("### Recall by slice (condition must hold in ground truth)")
        lines.append("")
        lines.append("| Slice | Recall | N |")
        lines.append("|---|---|---|")
        for cond, s in sorted(slices.items()):
            lines.append(f"| {cond} | {fmt(s['recall'])} | {s['n']} |")
    lines.append("")

    g = results["groups"]
    lines.append("## Grouping")
    if g.get("n", 0) >= 2:
        lines.append(f"- N labeled: **{g['n']}**")
        lines.append(f"- Adjusted Rand Index: **{fmt(g.get('ari'))}**")
        lines.append(f"- Pair precision: **{fmt(g.get('pair_precision'))}**")
        lines.append(f"- Pair recall: **{fmt(g.get('pair_recall'))}**")
        lines.append(f"- Pair F1: **{fmt(g.get('pair_f1'))}**")
        tier = g.get("tier_confusion", {})
        if tier:
            lines.append("")
            lines.append("### Expected link tier vs predicted group_type")
            lines.append("")
            lines.append("| Expected | Predicted | Count |")
            lines.append("|---|---|---|")
            for (exp, pred_t), c in sorted(tier.items()):
                lines.append(f"| {exp} | {pred_t} | {c} |")
    else:
        lines.append(f"- {g.get('note', 'not enough data')}")
    lines.append("")

    s = results["sharpness"]
    lines.append("## Focus / sharpness")
    if s.get("n"):
        lines.append(f"- N labeled: **{s['n']}**")
        lines.append(f"- Spearman ρ (human 1–10 vs raw sharpness): **{fmt(s.get('spearman_rho'))}**")
        lines.append(f"- MAE on 1–10 scale (after percentile mapping): **{fmt(s.get('mae_1_10'))}**")
        lines.append(f"- Category accuracy: **{fmt(s.get('category_accuracy'))}** "
                     f"(N={s.get('category_n', 0)})")
    else:
        lines.append(f"- {s.get('note', 'not enough data')}")

    lines.append("")
    lines.append("## Caveats")
    lines.append("")
    lines.append(
        "- Eye open/closed scores come from a **mock provider** (alternating 0/1) "
        "in the current build — the `eyes_open_count` and `eye_open` columns in the "
        "DB are not a real signal until the ONNX eye classifier ships."
    )
    lines.append(
        "- Sharpness percentile is computed **within this labeled sample**, not the "
        "full shoot. If you want to benchmark against the in-app percentiles, run "
        "the harness per-shoot with `--shoots <slug>`."
    )
    lines.append(
        "- Grouping ARI counts images not clustered by PhotoSift as singletons, so a "
        "systematic under-cluster shows up as recall loss rather than precision loss."
    )

    out_path.write_text("\n".join(lines), encoding="utf-8")


def fmt(v):
    if v is None or (isinstance(v, float) and math.isnan(v)):
        return "—"
    if isinstance(v, float):
        return f"{v:.3f}"
    return str(v)


# ----- Driver -----


def main():
    ap = argparse.ArgumentParser(description="Compute PhotoSift AI benchmark metrics.")
    ap.add_argument("--db", type=Path, default=Path.home() / ".photosift" / "photosift.db")
    ap.add_argument("--workbook", type=Path, default=Path(__file__).parent / "photosift_benchmark.xlsx")
    ap.add_argument("--report", type=Path, default=Path(__file__).parent / "benchmark_report.md")
    ap.add_argument("--shoots", nargs="*", help="Optional list of shoot slugs to restrict evaluation to.")
    args = ap.parse_args()

    if not args.db.exists():
        print(f"ERROR: DB not found at {args.db}", file=sys.stderr)
        return 2
    if not args.workbook.exists():
        print(f"ERROR: workbook not found at {args.workbook}", file=sys.stderr)
        return 2

    wb = load_workbook(args.workbook)

    # Collect all raw_paths from Faces sheet (superset — same set is on Groups/Sharp).
    faces_ws = wb["Faces"]
    raw_paths = set()
    for row in faces_ws.iter_rows(min_row=2, values_only=True):
        if row and row[1]:
            raw_paths.add(row[1])

    if not raw_paths:
        print("ERROR: no rows in Faces sheet — run sample_library.py first", file=sys.stderr)
        return 2

    loaded = load_db_rows(args.db, raw_paths)
    db_rows = loaded["rows"]
    settings = loaded["settings"]

    # Shoot filter — currently not enforced here because raw_path already
    # scopes us; exposed for future use when the harness takes a shoot-aware view.
    if args.shoots:
        print(f"(note: --shoots filter currently informational, all matched rows kept)")

    matched = len(db_rows)
    print(f"matched {matched}/{len(raw_paths)} labeled rows to PhotoSift DB")

    results = {
        "faces": eval_faces(faces_ws, db_rows),
        "groups": eval_groups(wb["Groups"], db_rows),
        "sharpness": eval_sharpness(wb["Sharpness"], db_rows),
    }

    meta = {
        "run_at": datetime.now().isoformat(timespec="seconds"),
        "db_path": str(args.db),
        "n_samples_total": len(raw_paths),
        "n_faces_labeled": results["faces"].get("n", 0),
        "n_groups_labeled": results["groups"].get("n", 0),
        "n_sharpness_labeled": results["sharpness"].get("n", 0),
        "near_dup_threshold": settings.get("near_dup_threshold"),
        "related_threshold": settings.get("related_threshold"),
    }

    write_results(wb, results, meta)
    wb.save(args.workbook)
    write_report(results, meta, args.report)

    # Also dump to stdout as JSON for scripting.
    print(json.dumps({"results": sanitize(results), "meta": sanitize(meta)}, indent=2))
    print(f"\nworkbook updated: {args.workbook}")
    print(f"report written:   {args.report}")
    return 0


def sanitize(obj):
    if isinstance(obj, dict):
        return {str(k): sanitize(v) for k, v in obj.items()}
    if isinstance(obj, (list, tuple)):
        return [sanitize(v) for v in obj]
    if isinstance(obj, float) and math.isnan(obj):
        return None
    return obj


if __name__ == "__main__":
    import sys
    sys.exit(main())
