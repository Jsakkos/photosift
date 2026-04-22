"""Build the PhotoSift benchmark labeling workbook from scratch.

Run once to produce `photosift_benchmark.xlsx`. The sampler
(`sample_library.py`) appends image rows into the per-feature sheets; the
harness (`run_benchmark.py`) reads those rows plus PhotoSift's SQLite to
compute metrics.

Intentionally simple: one sheet per feature, data-validation dropdowns on
the columns a human fills in, a cover sheet with links and the labeling
protocol in-line, and a `_RunMeta` sheet the harness writes to on each run.
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUT = Path(__file__).parent / "photosift_benchmark.xlsx"

FONT_NAME = "Arial"

HEADER_FILL = PatternFill("solid", start_color="1F2937")  # slate-800
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=11)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

SECTION_FILL = PatternFill("solid", start_color="E5E7EB")  # slate-200
SECTION_FONT = Font(name=FONT_NAME, bold=True, size=12)

INPUT_FILL = PatternFill("solid", start_color="FFFBEB")  # amber-50
COMPUTED_FILL = PatternFill("solid", start_color="F3F4F6")  # slate-100

THIN = Side(border_style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header(row):
    for c in row:
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = HEADER_ALIGN
        c.border = BORDER


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_dropdown(ws, col_letter, options, start_row=2, end_row=1000):
    values = ",".join(options)
    dv = DataValidation(type="list", formula1=f'"{values}"', allow_blank=True)
    dv.error = f"Must be one of: {', '.join(options)}"
    dv.errorTitle = "Invalid value"
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}{start_row}:{col_letter}{end_row}")


def add_overview(wb):
    ws = wb.create_sheet("Overview", 0)
    ws.sheet_view.showGridLines = False

    ws["A1"] = "PhotoSift AI Benchmark — Labeling Workbook"
    ws["A1"].font = Font(name=FONT_NAME, bold=True, size=18)
    ws["A2"] = (
        "Ground truth for three features: face detection, perceptual grouping, "
        "and focus/eye sharpness scoring. Populate the yellow columns; the harness "
        "fills the gray ones."
    )
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[2].height = 40

    # Sheet guide
    ws["A4"] = "Sheets in this workbook"
    ws["A4"].font = SECTION_FONT
    ws["A4"].fill = SECTION_FILL
    ws.merge_cells("A4:C4")

    guide = [
        ("Faces", "One row per image. Label the true face count + conditions."),
        ("Groups", "One row per image with a cluster_id you assign. Images sharing a cluster_id are expected to group."),
        ("Sharpness", "One row per image. Rate focus 1–10 and label category (tack / soft / motion / oof)."),
        ("_Samples", "All sampled images (populated by sample_library.py). Master list with EXIF."),
        ("_RunMeta", "Written by run_benchmark.py — thresholds used, dataset stats, timestamps."),
        ("Results", "Written by run_benchmark.py — per-feature metrics + per-slice breakdowns."),
    ]
    for i, (name, desc) in enumerate(guide, start=5):
        ws.cell(row=i, column=1, value=name).font = Font(name=FONT_NAME, bold=True)
        ws.cell(row=i, column=2, value=desc).alignment = Alignment(wrap_text=True)

    # Legend
    start = 5 + len(guide) + 1
    ws.cell(row=start, column=1, value="Cell legend").font = SECTION_FONT
    ws.cell(row=start, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=start, start_column=1, end_row=start, end_column=3)

    ws.cell(row=start + 1, column=1, value="Yellow = you fill in (ground truth)").fill = INPUT_FILL
    ws.cell(row=start + 2, column=1, value="Gray = harness fills (PhotoSift prediction)").fill = COMPUTED_FILL
    ws.cell(row=start + 3, column=1, value="White = reference (path, EXIF, notes)")

    # Labeling protocol
    p = start + 5
    ws.cell(row=p, column=1, value="Labeling protocol — read before you start").font = SECTION_FONT
    ws.cell(row=p, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=p, start_column=1, end_row=p, end_column=3)

    protocol = [
        "1. Run sample_library.py pointed at your RAW library. It populates Faces / Groups / Sharpness / _Samples.",
        "2. For FACES: count every face that a human would recognize as a face in a hypothetical 1500px-wide version of the image. Tiny background faces (<2% of frame height) count — PhotoSift should catch them. Sunglasses count; full profile counts; partial occlusion counts if >50% visible. Fill face_count_true. Mark conditions that apply (multi-select columns).",
        "3. For GROUPS: pick an integer cluster_id per image. Images you'd expect PhotoSift to cluster together get the same cluster_id. Unique shots get a unique cluster_id. For borderline cases (reframes, zooms), mark expected_link = near_dup or related to encode your expectation of the tier.",
        "4. For SHARPNESS: give focus_score_true from 1 (unusable) to 10 (tack sharp, critical focus nailed). Then pick ONE category: tack / good / acceptable_soft / motion_blur / oof / shake. Focus on the main subject — if there's no clear subject, score on whatever was intended to be sharp.",
        "5. Don't look at PhotoSift's predictions before labeling. Label blind. The harness compares your labels against the DB after the fact.",
        "6. If you're unsure on an image, mark unsure = Y and skip it. The harness will exclude unsure rows from metrics but keep them for qualitative review.",
    ]
    for i, line in enumerate(protocol):
        cell = ws.cell(row=p + 1 + i, column=1, value=line)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=p + 1 + i, start_column=1, end_row=p + 1 + i, end_column=3)
        ws.row_dimensions[p + 1 + i].height = 30

    # Focus score rubric — kept here rather than on the Sharpness sheet so
    # the sampler can safely wipe sample rows without destroying reference.
    r = p + 1 + len(protocol) + 1
    ws.cell(row=r, column=1, value="Focus score rubric (1–10 for Sharpness sheet)").font = SECTION_FONT
    ws.cell(row=r, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    rubric = [
        (10, "Tack sharp — critical focus nailed, eye-level on subject, no visible softness at 100%"),
        (9, "Very sharp — no visible issue, minor microcontrast gap vs 10"),
        (8, "Sharp — in focus, slightly soft at pixel-peep"),
        (7, "Acceptable — subject in focus but edges aren't crisp; prints fine"),
        (6, "Slightly soft — focus plane off subject a bit, or mild diffraction / camera shake"),
        (5, "Soft — usable for small web, not for print"),
        (4, "Noticeably soft — mild motion blur or focus miss"),
        (3, "Soft and blurry — cull unless the moment is irreplaceable"),
        (2, "Blurry — shake or focus miss, subject not recoverable"),
        (1, "Unusable — severe motion blur, OOF, shake"),
    ]
    for i, (score, desc) in enumerate(rubric, start=1):
        ws.cell(row=r + i, column=1, value=score).font = Font(name=FONT_NAME, bold=True)
        ws.cell(row=r + i, column=2, value=desc)
        ws.cell(row=r + i, column=2).alignment = Alignment(wrap_text=True)
        ws.merge_cells(start_row=r + i, start_column=2, end_row=r + i, end_column=3)

    set_col_widths(ws, [22, 70, 20])


def add_faces_sheet(wb):
    ws = wb.create_sheet("Faces")
    headers = [
        # identity
        ("sample_id", 10),
        ("raw_path", 55),
        ("filename", 22),
        # ground truth (YOU fill)
        ("face_count_true", 14),
        ("has_profile", 11),
        ("has_small_face", 13),
        ("has_occlusion", 13),
        ("has_sunglasses", 13),
        ("has_low_light", 13),
        ("has_motion", 11),
        ("has_backlit", 11),
        ("unsure", 9),
        ("notes", 30),
        # PhotoSift prediction (harness fills)
        ("face_count_pred", 14),
        ("max_confidence", 13),
        ("min_confidence", 13),
        ("error_type", 14),  # missed / extra / correct / unsure
        ("abs_error", 10),
    ]
    ws.append([h for h, _ in headers])
    style_header(ws[1])
    set_col_widths(ws, [w for _, w in headers])
    ws.freeze_panes = "D2"

    # Shade input vs computed columns (from row 2 down a generous range).
    input_cols = ["D", "E", "F", "G", "H", "I", "J", "K", "L", "M"]
    computed_cols = ["N", "O", "P", "Q", "R"]
    for col in input_cols:
        for r in range(2, 2001):
            ws[f"{col}{r}"].fill = INPUT_FILL
    for col in computed_cols:
        for r in range(2, 2001):
            ws[f"{col}{r}"].fill = COMPUTED_FILL

    yn = ["Y", "N"]
    for col in ["E", "F", "G", "H", "I", "J", "K", "L"]:
        add_dropdown(ws, col, yn, start_row=2, end_row=2000)

    add_dropdown(ws, "Q", ["missed", "extra", "correct", "unsure"], start_row=2, end_row=2000)


def add_groups_sheet(wb):
    ws = wb.create_sheet("Groups")
    headers = [
        ("sample_id", 10),
        ("raw_path", 55),
        ("filename", 22),
        ("capture_time", 20),
        # ground truth
        ("cluster_id_true", 14),
        ("expected_link", 14),  # near_dup / related / unique
        ("burst_type", 16),     # burst / reframe / bracket / panorama / timelapse / unique / other
        ("unsure", 9),
        ("notes", 30),
        # PhotoSift prediction
        ("cluster_id_pred", 14),
        ("group_type_pred", 16),
        ("min_hamming_intra", 17),  # min hamming to another same-cluster image
        ("min_hamming_inter", 17),  # min hamming to a different-cluster image
    ]
    ws.append([h for h, _ in headers])
    style_header(ws[1])
    set_col_widths(ws, [w for _, w in headers])
    ws.freeze_panes = "E2"

    input_cols = ["E", "F", "G", "H", "I"]
    computed_cols = ["J", "K", "L", "M"]
    for col in input_cols:
        for r in range(2, 2001):
            ws[f"{col}{r}"].fill = INPUT_FILL
    for col in computed_cols:
        for r in range(2, 2001):
            ws[f"{col}{r}"].fill = COMPUTED_FILL

    add_dropdown(ws, "F", ["near_dup", "related", "unique"], start_row=2, end_row=2000)
    add_dropdown(ws, "G", ["burst", "reframe", "bracket", "panorama", "timelapse", "unique", "other"], start_row=2, end_row=2000)
    add_dropdown(ws, "H", ["Y", "N"], start_row=2, end_row=2000)


def add_sharpness_sheet(wb):
    ws = wb.create_sheet("Sharpness")
    headers = [
        ("sample_id", 10),
        ("raw_path", 55),
        ("filename", 22),
        ("camera", 18),
        ("lens", 20),
        ("shutter_speed", 14),
        ("aperture", 10),
        ("iso", 8),
        # ground truth
        ("focus_score_true", 16),   # 1-10
        ("category_true", 18),      # tack / good / acceptable_soft / motion_blur / oof / shake
        ("subject_type", 14),       # face / eye / animal / landscape / object / mixed
        ("unsure", 9),
        ("notes", 30),
        # PhotoSift prediction
        ("sharpness_raw", 14),
        ("sharpness_pct", 14),      # percentile within the sampled shoot
        ("bucketed_pred", 14),      # tack / good / acceptable_soft / motion_blur / oof / shake derived by harness
        ("face_mean_eye_sharp", 18),
        ("eyes_open_count", 16),
        ("quality_score_db", 15),
        ("abs_score_delta", 14),
    ]
    ws.append([h for h, _ in headers])
    style_header(ws[1])
    set_col_widths(ws, [w for _, w in headers])
    ws.freeze_panes = "I2"

    input_cols = ["I", "J", "K", "L", "M"]
    computed_cols = ["N", "O", "P", "Q", "R", "S", "T"]
    for col in input_cols:
        for r in range(2, 2001):
            ws[f"{col}{r}"].fill = INPUT_FILL
    for col in computed_cols:
        for r in range(2, 2001):
            ws[f"{col}{r}"].fill = COMPUTED_FILL

    add_dropdown(ws, "J", ["tack", "good", "acceptable_soft", "motion_blur", "oof", "shake"], start_row=2, end_row=2000)
    add_dropdown(ws, "K", ["face", "eye", "animal", "landscape", "object", "mixed"], start_row=2, end_row=2000)
    add_dropdown(ws, "L", ["Y", "N"], start_row=2, end_row=2000)
    # Rubric lives in Overview (see add_overview) so the sampler can freely
    # wipe + repopulate this sheet without losing reference material.


def add_samples_sheet(wb):
    ws = wb.create_sheet("_Samples")
    headers = [
        ("sample_id", 10),
        ("raw_path", 55),
        ("filename", 22),
        ("ext", 6),
        ("shoot_slug", 22),
        ("stratum", 16),
        ("capture_time", 20),
        ("camera", 18),
        ("lens", 22),
        ("focal_length", 12),
        ("aperture", 10),
        ("shutter_speed", 13),
        ("iso", 8),
        ("width_px", 10),
        ("height_px", 10),
        ("file_size_mb", 12),
        ("content_hash", 18),
        ("photo_id_db", 12),
        ("ai_analyzed_at", 22),
    ]
    ws.append([h for h, _ in headers])
    style_header(ws[1])
    set_col_widths(ws, [w for _, w in headers])
    ws.freeze_panes = "B2"


def add_runmeta_sheet(wb):
    ws = wb.create_sheet("_RunMeta")
    ws["A1"] = "PhotoSift benchmark run metadata"
    ws["A1"].font = SECTION_FONT
    ws["A3"] = "key"
    ws["B3"] = "value"
    style_header(ws[3])
    set_col_widths(ws, [28, 60])
    # Rows will be filled by run_benchmark.py.
    defaults = [
        ("run_at", ""),
        ("db_path", ""),
        ("library_root", ""),
        ("n_samples_total", ""),
        ("n_faces_labeled", ""),
        ("n_groups_labeled", ""),
        ("n_sharpness_labeled", ""),
        ("near_dup_threshold", ""),
        ("related_threshold", ""),
        ("face_conf_threshold", ""),
        ("providerStatus", ""),
        ("eyeProviderKind", ""),
        ("mouthProviderKind", ""),
        ("notes", ""),
    ]
    for i, (k, v) in enumerate(defaults, start=4):
        ws.cell(row=i, column=1, value=k)
        ws.cell(row=i, column=2, value=v)


def add_results_sheet(wb):
    ws = wb.create_sheet("Results")
    ws["A1"] = "Results — written by run_benchmark.py"
    ws["A1"].font = SECTION_FONT
    ws["A2"] = "Don't edit by hand. Re-run the harness to refresh."
    ws["A2"].font = Font(name=FONT_NAME, italic=True, color="6B7280")
    set_col_widths(ws, [28, 18, 14, 14, 14, 60])
    ws.append([])
    ws.append(["metric", "feature", "value", "n", "ci_95", "notes"])
    style_header(ws[4])


def main():
    wb = Workbook()
    # Drop the default sheet; we add Overview as the first sheet explicitly.
    default = wb.active
    wb.remove(default)

    add_overview(wb)
    add_faces_sheet(wb)
    add_groups_sheet(wb)
    add_sharpness_sheet(wb)
    add_samples_sheet(wb)
    add_runmeta_sheet(wb)
    add_results_sheet(wb)

    # Apply Arial to every cell we wrote — cheaper as a pass than threading it
    # through every style call above.
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.font and cell.font.name != FONT_NAME:
                    existing = cell.font
                    cell.font = Font(
                        name=FONT_NAME,
                        bold=existing.bold,
                        italic=existing.italic,
                        size=existing.size or 11,
                        color=existing.color,
                    )

    wb.save(OUT)
    print(f"wrote {OUT}")


if __name__ == "__main__":
    main()
