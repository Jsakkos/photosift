"""End-to-end smoke test of the harness against a synthetic DB + labeled
workbook. Not shipped to users — just verifies the harness works before the
user has a real DB to point it at."""

from __future__ import annotations

import shutil
import sqlite3
import subprocess
import sys
from pathlib import Path

from openpyxl import load_workbook

import os

HERE = Path(__file__).parent
WB = HERE / "photosift_benchmark.xlsx"
# SQLite journaling needs a filesystem that supports it; the session mount
# sometimes rejects lock files. Allow an override for CI/dev.
FIXTURE_DIR = Path(os.environ.get("PHOTOSIFT_BENCH_FIXTURE", HERE / "_smoke_fixture"))
FIXTURE_DIR.mkdir(parents=True, exist_ok=True)
TEST_WB = FIXTURE_DIR / "photosift_benchmark.xlsx"
TEST_DB = FIXTURE_DIR / "photosift.db"


def build_fixture_db(db_path: Path, rows: list[dict]) -> None:
    """Mirror the prod schema closely enough that the harness queries land.
    We only create the columns the harness reads."""
    if db_path.exists():
        db_path.unlink()
    con = sqlite3.connect(str(db_path))
    con.executescript("""
    CREATE TABLE shoots (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        slug TEXT, date TEXT, source_path TEXT, dest_path TEXT,
        photo_count INTEGER DEFAULT 0,
        imported_at TEXT DEFAULT (datetime('now')),
        import_mode TEXT DEFAULT 'copy'
    );
    CREATE TABLE photos (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        shoot_id INTEGER, filename TEXT, raw_path TEXT, preview_path TEXT,
        thumb_path TEXT, content_hash BLOB, phash BLOB,
        exif_date TEXT, camera TEXT, lens TEXT,
        focal_length REAL, aperture REAL, shutter_speed TEXT, iso INTEGER,
        orientation INTEGER,
        flag TEXT DEFAULT 'unreviewed', destination TEXT DEFAULT 'unrouted',
        star_rating INTEGER DEFAULT 0,
        sharpness_score REAL, quality_score REAL,
        face_count INTEGER, eyes_open_count INTEGER, ai_analyzed_at TEXT
    );
    CREATE TABLE faces (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        photo_id INTEGER,
        bbox_x REAL, bbox_y REAL, bbox_w REAL, bbox_h REAL,
        left_eye_x REAL, left_eye_y REAL, right_eye_x REAL, right_eye_y REAL,
        left_eye_open INTEGER, right_eye_open INTEGER,
        left_eye_sharpness REAL, right_eye_sharpness REAL,
        detection_confidence REAL, smile_score REAL
    );
    CREATE TABLE groups (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        shoot_id INTEGER, group_type TEXT
    );
    CREATE TABLE group_members (
        group_id INTEGER, photo_id INTEGER, is_cover INTEGER DEFAULT 0,
        PRIMARY KEY(group_id, photo_id)
    );
    CREATE TABLE settings (
        id INTEGER PRIMARY KEY CHECK (id = 1),
        near_dup_threshold INTEGER DEFAULT 4,
        related_threshold INTEGER DEFAULT 12
    );
    INSERT INTO settings (id) VALUES (1);
    INSERT INTO shoots (slug, date, source_path, dest_path) VALUES ('fixture', '2026-01-01', '/src', '/dst');
    """)

    cur = con.cursor()
    for i, r in enumerate(rows, start=1):
        cur.execute("""
            INSERT INTO photos (id, shoot_id, filename, raw_path, preview_path, thumb_path,
                                content_hash, phash, face_count, eyes_open_count,
                                sharpness_score, quality_score, ai_analyzed_at)
            VALUES (?, 1, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, datetime('now'))
        """, (
            i, Path(r["raw_path"]).name, r["raw_path"], f"/preview/{i}.jpg", f"/thumb/{i}.jpg",
            b"h" * 32, r.get("phash"), r.get("face_count"), r.get("eyes_open_count"),
            r.get("sharpness_score"), r.get("quality_score"),
        ))
        for f in r.get("faces", []):
            cur.execute("""
                INSERT INTO faces (photo_id, bbox_x, bbox_y, bbox_w, bbox_h,
                                   left_eye_x, left_eye_y, right_eye_x, right_eye_y,
                                   left_eye_open, right_eye_open,
                                   left_eye_sharpness, right_eye_sharpness,
                                   detection_confidence)
                VALUES (?, 0.4, 0.3, 0.2, 0.2, 0.45, 0.35, 0.55, 0.35,
                        ?, ?, ?, ?, ?)
            """, (i, f.get("le", 1), f.get("re", 1), f.get("ls", 80.0), f.get("rs", 80.0), f.get("conf", 0.85)))
        if r.get("group_id"):
            gid = r["group_id"]
            cur.execute("INSERT OR IGNORE INTO groups (id, shoot_id, group_type) VALUES (?, 1, ?)",
                        (gid, r.get("group_type", "near_duplicate")))
            cur.execute("INSERT INTO group_members (group_id, photo_id) VALUES (?, ?)", (gid, i))
    con.commit()
    con.close()


def build_labeled_workbook():
    shutil.copy(WB, TEST_WB)
    wb = load_workbook(TEST_WB)

    # Design the labels. 6 images; design known-outcome predictions in DB.
    labels = [
        # raw_path                 truth_faces  slice_flags          cluster  link      focus  cat
        ("/lib/img001.nef",        1,  {"has_profile":"N"},           10, "near_dup", 9, "tack"),
        ("/lib/img002.nef",        1,  {"has_profile":"N"},           10, "near_dup", 9, "tack"),
        ("/lib/img003.nef",        2,  {"has_small_face":"Y"},        11, "related",  7, "good"),
        ("/lib/img004.nef",        0,  {},                            12, "unique",   4, "motion_blur"),
        ("/lib/img005.nef",        3,  {"has_occlusion":"Y"},         13, "unique",   8, "good"),
        ("/lib/img006.nef",        1,  {},                            10, "near_dup", 2, "oof"),
    ]

    # Populate Faces sheet.
    ws = wb["Faces"]
    headers = {c.value: i for i, c in enumerate(ws[1])}
    # Ensure no lingering rows from first build.
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    for sid, (raw, fc, flags, _c, _l, _fs, _cat) in enumerate(labels, start=1):
        row = [None] * ws.max_column
        row[headers["sample_id"]] = sid
        row[headers["raw_path"]] = raw
        row[headers["filename"]] = Path(raw).name
        row[headers["face_count_true"]] = fc
        for k, v in flags.items():
            row[headers[k]] = v
        ws.append(row)

    # Populate Groups sheet.
    ws = wb["Groups"]
    headers = {c.value: i for i, c in enumerate(ws[1])}
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    for sid, (raw, _fc, _flags, cluster, link, _fs, _cat) in enumerate(labels, start=1):
        row = [None] * ws.max_column
        row[headers["sample_id"]] = sid
        row[headers["raw_path"]] = raw
        row[headers["filename"]] = Path(raw).name
        row[headers["cluster_id_true"]] = cluster
        row[headers["expected_link"]] = link
        ws.append(row)

    # Populate Sharpness sheet.
    ws = wb["Sharpness"]
    headers = {c.value: i for i, c in enumerate(ws[1])}
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    for sid, (raw, _fc, _flags, _c, _l, fs, cat) in enumerate(labels, start=1):
        row = [None] * ws.max_column
        row[headers["sample_id"]] = sid
        row[headers["raw_path"]] = raw
        row[headers["filename"]] = Path(raw).name
        row[headers["focus_score_true"]] = fs
        row[headers["category_true"]] = cat
        ws.append(row)

    wb.save(TEST_WB)
    return labels


def build_db_rows(labels):
    # Design predictions so we can verify the metrics are sane:
    #   img001: correct 1 face, same cluster 10 as img002 & img006
    #   img002: correct 1 face
    #   img003: MISS — DB says 1 face, truth is 2
    #   img004: correct 0 faces
    #   img005: EXTRA — DB says 4 faces (overcount)
    #   img006: correct 1 face; but DB groups it separately (miss intra-cluster)
    rows = []
    # Provide hex phashes that roughly correlate with cluster membership.
    phash_by_cluster = {
        10: bytes.fromhex("00" * 8),
        11: bytes.fromhex("0f" * 8),
        12: bytes.fromhex("f0" * 8),
        13: bytes.fromhex("ff" * 8),
    }
    pred_face_counts = [1, 1, 1, 0, 4, 1]
    pred_sharpness = [95.0, 92.0, 70.0, 30.0, 80.0, 10.0]
    pred_clusters = [100, 100, 101, None, None, None]  # img006 not grouped → miss

    for i, (raw, fc, _flags, cluster, _l, _fs, _cat) in enumerate(labels):
        pc = pred_face_counts[i]
        rows.append({
            "raw_path": raw,
            "face_count": pc,
            "eyes_open_count": pc,
            "sharpness_score": pred_sharpness[i],
            "quality_score": pred_sharpness[i] * 0.85,
            "phash": phash_by_cluster[cluster],
            "faces": [{} for _ in range(pc)],
            "group_id": pred_clusters[i],
            "group_type": "near_duplicate" if pred_clusters[i] == 100 else "related",
        })
    return rows


def main():
    print("building fixture DB...")
    labels = build_labeled_workbook()
    rows = build_db_rows(labels)
    build_fixture_db(TEST_DB, rows)

    print("running harness...")
    proc = subprocess.run(
        [sys.executable, "run_benchmark.py",
         "--db", str(TEST_DB),
         "--workbook", str(TEST_WB),
         "--report", str(FIXTURE_DIR / "benchmark_report.md")],
        cwd=HERE, capture_output=True, text=True,
    )
    print("stdout:")
    print(proc.stdout)
    print("stderr:")
    print(proc.stderr)
    if proc.returncode != 0:
        sys.exit(proc.returncode)

    # Sanity checks.
    wb = load_workbook(TEST_WB)
    ws = wb["Faces"]
    headers = {c.value: i for i, c in enumerate(ws[1])}
    for row in ws.iter_rows(min_row=2, values_only=True):
        sid, raw, fn = row[0], row[1], row[2]
        et = row[headers["error_type"]]
        print(f"  sample {sid} {fn}: error_type={et}, pred={row[headers['face_count_pred']]}, true={row[headers['face_count_true']]}")


if __name__ == "__main__":
    main()
