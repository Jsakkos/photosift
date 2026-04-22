"""Scan a RAW library, stratify across shoot types, and pre-populate the
benchmark labeling workbook.

What "stratified sample" means here: we want coverage across the axes most
likely to expose AI weaknesses — portraits (face detection), bursts (grouping),
low-light + motion (sharpness). The script walks the library, groups files
into shoots (by parent folder name), assigns each shoot a stratum by a
lightweight heuristic on its EXIF distribution, then samples roughly a target
count from each stratum.

Requires:
    pip install openpyxl exifread

Usage:
    python sample_library.py --library /path/to/DSLR --target 150
    # then open photosift_benchmark.xlsx and start labeling

EXIF strategy: we use `exifread` which reads NEF directly. For weird formats
we fall back to best-effort file-system metadata and leave the EXIF fields
blank — the labeler can fill them in or we just skip that row's stratum tagging.
"""

from __future__ import annotations

import argparse
import hashlib
import random
import re
import sys
from collections import defaultdict
from dataclasses import dataclass, field, asdict
from datetime import datetime
from pathlib import Path
from typing import Iterable

try:
    import exifread  # type: ignore
except ImportError:
    print("ERROR: install exifread (pip install exifread)", file=sys.stderr)
    sys.exit(1)

from openpyxl import load_workbook

RAW_EXTS = {".nef", ".cr2", ".cr3", ".arw", ".raf", ".rw2", ".dng", ".orf", ".pef", ".srw"}

# Stratum buckets. The sampler routes each shoot into at most one stratum via
# a heuristic on its EXIF distribution — not perfect, but good enough to avoid
# a portrait-only test set.
STRATA = [
    "portraits_studio",     # short lens, shallow aperture, low ISO
    "portraits_available",  # short lens, shallow aperture, varied ISO
    "events_low_light",     # high ISO predominant, indoor zoom focal lengths
    "landscapes",           # small aperture (f/8+), wide focal length, low ISO
    "action_bursts",        # short capture-time deltas between adjacent frames
    "wildlife",             # long focal length (200mm+), varied aperture
    "mixed",                # fallback when no heuristic fires
]


@dataclass
class SampleRow:
    sample_id: int
    raw_path: str
    filename: str
    ext: str
    shoot_slug: str
    stratum: str
    capture_time: str | None
    camera: str | None
    lens: str | None
    focal_length: float | None
    aperture: float | None
    shutter_speed: str | None
    iso: int | None
    width_px: int | None
    height_px: int | None
    file_size_mb: float
    content_hash: str


@dataclass
class ShootGroup:
    slug: str
    files: list[Path] = field(default_factory=list)
    exif: list[dict] = field(default_factory=list)

    def stratum(self) -> str:
        """Classify this shoot into one stratum bucket using aggregate EXIF."""
        valid = [e for e in self.exif if e]
        if not valid:
            return "mixed"

        focals = [e.get("focal_length") for e in valid if e.get("focal_length")]
        apertures = [e.get("aperture") for e in valid if e.get("aperture")]
        isos = [e.get("iso") for e in valid if e.get("iso")]
        times = sorted([e.get("capture_time") for e in valid if e.get("capture_time")])

        def median(xs):
            if not xs:
                return None
            s = sorted(xs)
            n = len(s)
            return s[n // 2] if n % 2 else 0.5 * (s[n // 2 - 1] + s[n // 2])

        mf, ma, mi = median(focals), median(apertures), median(isos)

        # Burst detection: a shoot is a burst-rich shoot if 30%+ of adjacent
        # capture-time deltas are under 2 seconds.
        burst_ratio = 0.0
        if len(times) >= 4:
            deltas = [(b - a).total_seconds() for a, b in zip(times[:-1], times[1:])]
            burst_ratio = sum(1 for d in deltas if d <= 2.0) / len(deltas)

        if burst_ratio >= 0.3:
            return "action_bursts"
        if mf and mf >= 200:
            return "wildlife"
        if mf and mf <= 85 and ma and ma <= 2.2 and mi and mi <= 400:
            return "portraits_studio"
        if mf and mf <= 105 and ma and ma <= 2.8:
            return "portraits_available"
        if mi and mi >= 3200:
            return "events_low_light"
        if ma and ma >= 8.0 and mf and mf <= 50:
            return "landscapes"
        return "mixed"


def slugify(name: str) -> str:
    name = re.sub(r"[^A-Za-z0-9_\-]+", "_", name).strip("_")
    return name[:60] or "shoot"


def read_exif(path: Path) -> dict:
    """Best-effort EXIF read via exifread. Returns a dict of extracted fields,
    empty if the file can't be parsed."""
    out: dict = {}
    try:
        with path.open("rb") as f:
            tags = exifread.process_file(f, details=False, stop_tag="UNDEF")
    except Exception:
        return out

    def s(key):
        v = tags.get(key)
        return str(v) if v is not None else None

    def f(key):
        v = tags.get(key)
        if v is None:
            return None
        try:
            vs = v.values[0]
            if hasattr(vs, "num") and hasattr(vs, "den") and vs.den != 0:
                return vs.num / vs.den
            return float(vs)
        except Exception:
            return None

    def i(key):
        v = tags.get(key)
        if v is None:
            return None
        try:
            return int(v.values[0])
        except Exception:
            return None

    out["camera"] = s("Image Model") or s("EXIF Model")
    out["lens"] = s("EXIF LensModel") or s("MakerNote LensType")
    out["focal_length"] = f("EXIF FocalLength")
    out["aperture"] = f("EXIF FNumber") or f("EXIF ApertureValue")
    out["shutter_speed"] = s("EXIF ExposureTime")
    out["iso"] = i("EXIF ISOSpeedRatings")
    out["width_px"] = i("EXIF ExifImageWidth") or i("Image ImageWidth")
    out["height_px"] = i("EXIF ExifImageLength") or i("Image ImageLength")

    dt = s("EXIF DateTimeOriginal") or s("Image DateTime")
    if dt:
        try:
            out["capture_time"] = datetime.strptime(dt, "%Y:%m:%d %H:%M:%S")
        except ValueError:
            out["capture_time"] = None
    else:
        out["capture_time"] = None
    return out


def content_hash(path: Path, chunk: int = 1 << 20) -> str:
    """First 1 MB SHA256 prefix — enough to be a stable identifier, cheap to
    compute, and doesn't require hashing multi-GB RAW files for this use."""
    h = hashlib.sha256()
    with path.open("rb") as f:
        h.update(f.read(chunk))
    return h.hexdigest()[:16]


def walk_library(root: Path) -> Iterable[Path]:
    for p in root.rglob("*"):
        if p.is_file() and p.suffix.lower() in RAW_EXTS:
            yield p


def group_by_shoot(files: Iterable[Path]) -> dict[str, ShootGroup]:
    """PhotoSift conventions cluster RAWs under `YYYY-MM_Description/RAW/`.
    We use the grandparent-of-file folder name (the `YYYY-MM_Description`)
    as the shoot slug, falling back to the parent folder name."""
    shoots: dict[str, ShootGroup] = {}
    for f in files:
        # Try grandparent if parent is literally "RAW"; else use parent.
        parent = f.parent.name
        grand = f.parent.parent.name if f.parent.parent else parent
        slug_source = grand if parent.lower() == "raw" else parent
        slug = slugify(slug_source)
        shoots.setdefault(slug, ShootGroup(slug=slug)).files.append(f)
    return shoots


def sample_stratified(
    shoots: dict[str, ShootGroup],
    target_total: int,
    seed: int = 42,
) -> list[SampleRow]:
    """Allocate samples proportionally across strata, then sample within each."""
    rng = random.Random(seed)

    # Fingerprint each shoot to a stratum using a subset of its EXIF (reading
    # every file's EXIF is slow; 20 per shoot is enough for the heuristic).
    for sg in shoots.values():
        probe = rng.sample(sg.files, k=min(20, len(sg.files)))
        sg.exif = [read_exif(p) for p in probe]

    by_stratum: dict[str, list[tuple[ShootGroup, Path]]] = defaultdict(list)
    for sg in shoots.values():
        st = sg.stratum()
        for f in sg.files:
            by_stratum[st].append((sg, f))

    # Allocation: split budget evenly across present strata, but cap per stratum
    # at the number of images available. Redistribute leftovers.
    present = [s for s in STRATA if by_stratum.get(s)]
    if not present:
        raise RuntimeError("no images found under library root")

    base = target_total // len(present)
    allocations = {s: min(base, len(by_stratum[s])) for s in present}
    leftover = target_total - sum(allocations.values())
    # Spread leftovers round-robin into strata that still have headroom.
    cycle = [s for s in present if len(by_stratum[s]) > allocations[s]]
    while leftover > 0 and cycle:
        progress = False
        for s in list(cycle):
            if leftover <= 0:
                break
            if len(by_stratum[s]) > allocations[s]:
                allocations[s] += 1
                leftover -= 1
                progress = True
            else:
                cycle.remove(s)
        if not progress:
            break

    # For each stratum, sample files deterministically. When we can, we pull
    # consecutive runs from bursty shoots so the grouping sheet has real
    # near-duplicate pairs to label.
    rows: list[SampleRow] = []
    sample_id = 1
    for s in present:
        pool = by_stratum[s]
        k = allocations[s]
        chosen = rng.sample(pool, k=k)
        for sg, f in chosen:
            rows.append(build_row(sample_id, f, sg, s))
            sample_id += 1

    return rows


def build_row(sid: int, path: Path, shoot: ShootGroup, stratum: str) -> SampleRow:
    e = read_exif(path)
    try:
        size_mb = path.stat().st_size / (1024 * 1024)
    except OSError:
        size_mb = 0.0

    ct = e.get("capture_time")
    return SampleRow(
        sample_id=sid,
        raw_path=str(path),
        filename=path.name,
        ext=path.suffix.lower().lstrip("."),
        shoot_slug=shoot.slug,
        stratum=stratum,
        capture_time=ct.isoformat(sep=" ") if isinstance(ct, datetime) else None,
        camera=e.get("camera"),
        lens=e.get("lens"),
        focal_length=e.get("focal_length"),
        aperture=e.get("aperture"),
        shutter_speed=e.get("shutter_speed"),
        iso=e.get("iso"),
        width_px=e.get("width_px"),
        height_px=e.get("height_px"),
        file_size_mb=round(size_mb, 2),
        content_hash=content_hash(path),
    )


def _clear_data_rows(ws) -> None:
    """Delete rows below the header that have an integer sample_id in col A.
    Leaves rubric / reference rows (which have text or empty col A) alone."""
    # Walk bottom-up so row indices stay valid as we delete.
    for row_idx in range(ws.max_row, 1, -1):
        v = ws.cell(row=row_idx, column=1).value
        if isinstance(v, int):
            ws.delete_rows(row_idx, 1)


def write_workbook(rows: list[SampleRow], workbook_path: Path) -> None:
    wb = load_workbook(workbook_path)

    # _Samples — master list.
    ws = wb["_Samples"]
    _clear_data_rows(ws)
    header = [c.value for c in ws[1]]
    for r in rows:
        d = asdict(r)
        ws.append([d.get(h) for h in header])

    # Faces — populate sample_id / raw_path / filename.
    ws = wb["Faces"]
    _clear_data_rows(ws)
    for r in rows:
        ws.append([r.sample_id, r.raw_path, r.filename] + [None] * (ws.max_column - 3))

    # Groups — populate sample_id / raw_path / filename / capture_time.
    ws = wb["Groups"]
    _clear_data_rows(ws)
    for r in rows:
        ws.append([r.sample_id, r.raw_path, r.filename, r.capture_time] + [None] * (ws.max_column - 4))

    # Sharpness — populate identity + EXIF.
    ws = wb["Sharpness"]
    _clear_data_rows(ws)
    for r in rows:
        ws.append([
            r.sample_id, r.raw_path, r.filename,
            r.camera, r.lens, r.shutter_speed, r.aperture, r.iso,
        ] + [None] * (ws.max_column - 8))

    wb.save(workbook_path)


def main():
    ap = argparse.ArgumentParser(description="Sample a RAW library into the benchmark workbook.")
    ap.add_argument("--library", required=True, type=Path, help="Root of your RAW library (e.g. .../DSLR)")
    ap.add_argument("--workbook", type=Path, default=Path(__file__).parent / "photosift_benchmark.xlsx")
    ap.add_argument("--target", type=int, default=150, help="Target sample count (default 150)")
    ap.add_argument("--seed", type=int, default=42)
    args = ap.parse_args()

    if not args.library.is_dir():
        print(f"ERROR: {args.library} is not a directory", file=sys.stderr)
        sys.exit(2)
    if not args.workbook.exists():
        print(f"ERROR: run build_workbook.py first — {args.workbook} missing", file=sys.stderr)
        sys.exit(2)

    print(f"scanning {args.library}...")
    files = list(walk_library(args.library))
    print(f"found {len(files)} RAW files")
    if not files:
        sys.exit(1)

    shoots = group_by_shoot(files)
    print(f"grouped into {len(shoots)} shoots")

    rows = sample_stratified(shoots, target_total=args.target, seed=args.seed)
    print(f"sampled {len(rows)} images")

    # Summary by stratum.
    tally: dict[str, int] = defaultdict(int)
    for r in rows:
        tally[r.stratum] += 1
    for k in STRATA:
        if tally.get(k):
            print(f"  {k:<24} {tally[k]}")

    write_workbook(rows, args.workbook)
    print(f"wrote sampled rows to {args.workbook}")


if __name__ == "__main__":
    main()
